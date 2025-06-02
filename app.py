import streamlit as st
import openpyxl
from io import BytesIO
from datetime import datetime

st.title("Excel 做賬模板處理工具")

# ✅ 排序用函式（數字優先，其次字串）
def safe_sort_key(x):
    val = x[0] if len(x) > 0 else ""
    try:
        return (0, float(val))
    except (TypeError, ValueError):
        return (1, str(val) if val is not None else "")

uploaded_file = st.file_uploader("請上傳 Excel 檔案（.xlsx）", type="xlsx")

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file)

    sheetnames = wb.sheetnames

    # ✅ 第一分頁：根據 A 欄排序
    if len(sheetnames) >= 1:
        ws1 = wb[sheetnames[0]]
        rows = list(ws1.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data_rows = sorted(
                [r for r in rows[1:] if r and len(r) > 0],
                key=safe_sort_key
            )
            ws1.delete_rows(1, ws1.max_row)
            for row in [header] + data_rows:
                ws1.append(row)

    # ✅ 第二分頁：刪除 A、D、E 欄 → 排序
    if len(sheetnames) >= 2:
        ws2 = wb[sheetnames[1]]

        # 倒序刪除 E → D → A
        ws2.delete_cols(5)  # E 欄
        ws2.delete_cols(4)  # D 欄
        ws2.delete_cols(1)  # A 欄（原 A）

        # 排序
        rows = list(ws2.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data_rows = sorted(
                [r for r in rows[1:] if r and len(r) > 0],
                key=safe_sort_key
            )
            ws2.delete_rows(1, ws2.max_row)
            for row in [header] + data_rows:
                ws2.append(row)

    # ✅ 第三分頁整個刪除
    if len(sheetnames) > 2:
        wb.remove(wb[sheetnames[2]])
    sheetnames = wb.sheetnames

    # ✅ 第四分頁起清空內容（保留欄位結構）
    for sheet_name in sheetnames[3:]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value not in (None, ""):
                    cell.value = ""

    # ✅ 設定第一頁為啟動時顯示
    wb.active = 0

    # ✅ 檔名使用「上個月」為名稱
    now = datetime.now()
    month = now.month - 1 if now.month > 1 else 12

    try:
        dt = datetime.strptime(uploaded_file.name[:10], "%Y-%m-%d")
        parsed_month = dt.month - 1 if dt.month > 1 else 12
        month = parsed_month
    except:
        pass

    result_filename = f"{month}月做賬模板.xlsx"

    # ✅ 儲存檔案到記憶體
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("✅ 處理完成，請下載：")
    st.download_button(
        label="📥 下載做賬模板",
        data=output,
        file_name=result_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
